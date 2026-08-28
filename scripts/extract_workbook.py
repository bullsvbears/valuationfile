#!/usr/bin/env python3
"""
One-time importer: Software Valuation File (.xlsx) -> the dashboard's data tiers.

The workbook stored three kinds of numbers in visually identical cells:

  =FDS(..,"FE_ESTIMATE(..)")   a FactSet consensus estimate
  1234.5                       a hand-typed number (a reported actual or a fix)
  =[6]IS!$AY$6                 a link into the analyst's own model workbook

This script reads the cell *formulas*, not just their values, to recover which
of the three each number was, and writes them to the matching tier. That is the
provenance the spreadsheet lost the moment a value was entered.

Usage:  python3 scripts/extract_workbook.py <workbook.xlsx> [--out data]
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

# Column spans on the `Data` sheet, one contiguous block of years per metric.
METRIC_BLOCKS = {
    "revenue": (4, 14),
    "grossProfit": (16, 25),
    "ebitda": (27, 36),
    "eps": (38, 47),
    "fcf": (49, 58),
}
BALANCE_COLUMNS = {"shares": 60, "cash": 61, "debt": 62}

DATA_LAST_ROW = 352
DATA_LAST_COL = 62
MASTER_LAST_ROW = 444

EXTERNAL_LINK = re.compile(r"\[\d+\]")
FACTSET_MARKERS = ("FE_ESTIMATE", "FDS(", "_xll.")


def formula_text(cell):
    """Formula source for a cell, unwrapping openpyxl's array-formula objects."""
    value = cell.value
    return value.text if hasattr(value, "text") else value


def classify(formula) -> str:
    """Map a cell's formula to the tier that should own its value."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return "override"  # a hand-typed number
    if any(marker in formula for marker in FACTSET_MARKERS):
        return "factset"
    if EXTERNAL_LINK.search(formula):
        return "model"  # link into the analyst's own model workbook
    return "factset"  # derived in-sheet from FactSet-fed cells


def number_or_none(value):
    return value if isinstance(value, (int, float)) and not isinstance(value, bool) else None


def read_groups(workbook, sheet_name: str) -> dict[str, list[str]]:
    """Group memberships from the grouping sheets: a header row, then tickers."""
    sheet = workbook[sheet_name]
    groups: dict[str, list[str]] = {}
    current: str | None = None
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=3, values_only=True):
        label, ticker = row[1], row[2]
        if label is None:
            continue
        label = str(label).strip()
        if label in ("Mean", "Median"):
            continue
        if ticker is None:  # a group header carries no ticker
            current = label
            groups.setdefault(current, [])
            continue
        if current and isinstance(ticker, str) and ticker.strip():
            groups[current].append(ticker.strip())
    return groups


def read_prices(workbook) -> dict[str, dict]:
    """Latest price and returns from the Master Software sheet."""
    sheet = workbook["Master Software"]
    prices: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=6, max_row=MASTER_LAST_ROW, max_col=6, values_only=True):
        ticker = row[2]
        if not isinstance(ticker, str) or not ticker.strip():
            continue
        prices[ticker.strip()] = {
            "price": number_or_none(row[3]),
            "ytdReturn": number_or_none(row[4]),
            "priorYearReturn": number_or_none(row[5]),
        }
    return prices


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, default=Path("data"))
    args = parser.parse_args()

    values = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    formulas = openpyxl.load_workbook(args.workbook, data_only=False, read_only=True)

    value_rows = list(
        values["Data"].iter_rows(min_row=1, max_row=DATA_LAST_ROW, max_col=DATA_LAST_COL, values_only=True)
    )
    formula_rows = list(
        formulas["Data"].iter_rows(min_row=1, max_row=DATA_LAST_ROW, max_col=DATA_LAST_COL)
    )
    year_header = value_rows[2]  # row 3 carries the calendar year for each column

    sectors = read_groups(values, "Software Groups by Sector")
    peer_groups = read_groups(values, "Software Groups by Financials")
    prices = read_prices(values)

    sectors_by_ticker: dict[str, list[str]] = {}
    for name, members in sectors.items():
        for ticker in members:
            sectors_by_ticker.setdefault(ticker, []).append(name)
    peers_by_ticker: dict[str, list[str]] = {}
    for name, members in peer_groups.items():
        for ticker in members:
            peers_by_ticker.setdefault(ticker, []).append(name)

    companies: list[dict] = []
    seen: set[str] = set()
    duplicates: list[str] = []
    factset: dict[str, dict] = {}
    overrides: dict[str, dict] = {}
    models: dict[str, dict] = {}
    coverage: str | None = None

    def bucket(tier: str, ticker: str) -> dict:
        store = {"factset": factset, "override": overrides, "model": models}[tier]
        entry = store.setdefault(ticker, {"series": {}, "balance": {}})
        return entry

    for index in range(3, DATA_LAST_ROW):
        value_row, formula_row = value_rows[index], formula_rows[index]
        ticker, name = value_row[0], value_row[1]
        if ticker is None and name is None:
            continue
        if name is None:  # a section header such as "Bhatia - Covered Companies"
            coverage = str(ticker)
            continue

        ticker = str(ticker).strip()
        if ticker in seen:
            # The workbook looked companies up with VLOOKUP, which only ever
            # reaches the first matching row, so a repeated ticker was already
            # dead weight there. Carrying it forward would give two rows the
            # same identity.
            duplicates.append(f"{ticker} (row {index + 1})")
            continue
        seen.add(ticker)
        covered = bool(coverage and "Covered" in coverage and "Non-Covered" not in coverage)

        for metric, (first_col, last_col) in METRIC_BLOCKS.items():
            for col in range(first_col, last_col + 1):
                year = year_header[col - 1]
                value = number_or_none(value_row[col - 1])
                if year is None or value is None:
                    continue
                tier = classify(formula_text(formula_row[col - 1]))
                bucket(tier, ticker)["series"].setdefault(metric, {})[str(year)] = value

        for key, col in BALANCE_COLUMNS.items():
            value = number_or_none(value_row[col - 1])
            if value is None:
                continue
            tier = classify(formula_text(formula_row[col - 1]))
            bucket(tier, ticker)["balance"][key] = value

        quote = prices.get(ticker, {})
        if quote.get("price") is not None:
            bucket("factset", ticker)["price"] = quote["price"]

        companies.append(
            {
                "ticker": ticker,
                "name": str(name).strip(),
                "fiscalYearEnd": number_or_none(value_row[2]),
                "coverage": coverage,
                "covered": covered,
                "sectors": sectors_by_ticker.get(ticker, []),
                "peerGroups": peers_by_ticker.get(ticker, []),
                "ytdReturn": quote.get("ytdReturn"),
                "priorYearReturn": quote.get("priorYearReturn"),
            }
        )

    args.out.mkdir(parents=True, exist_ok=True)
    (args.out / "models").mkdir(exist_ok=True)

    write(args.out / "universe.json", {
        "companies": companies,
        "sectors": sectors,
        "peerGroups": peer_groups,
    })
    write(args.out / "factset-cache.json", {
        "asOf": None,
        "source": f"Imported from {args.workbook.name} (cached FDS values, not a live pull)",
        "companies": factset,
    })

    # Every override produced by this import is marked cell by cell, so an
    # analyst clearing their own edits never disturbs the imported history.
    imported_note = f"Hard-coded cell imported from {args.workbook.name}"
    write(args.out / "overrides.json", {
        "companies": {
            ticker: {
                **entry,
                "imported": {
                    metric: sorted(years, key=int) for metric, years in entry["series"].items()
                },
                "importNote": imported_note,
            }
            for ticker, entry in overrides.items()
        }
    })

    for ticker, entry in models.items():
        write(args.out / "models" / f"{ticker}.json", {"ticker": ticker, **entry})

    print(f"companies        {len(companies)}")
    if duplicates:
        print(f"skipped repeats  {', '.join(duplicates)}")
    print(f"factset tier     {len(factset)} companies")
    print(f"override tier    {len(overrides)} companies")
    print(f"model tier       {len(models)} companies")


def write(path: Path, payload) -> None:
    path.write_text(json.dumps(payload, indent=1, sort_keys=False) + "\n")


if __name__ == "__main__":
    main()
